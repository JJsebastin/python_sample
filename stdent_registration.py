from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
import pathlib

background = '#06283D'
framebg = '#EDEDED'
framefg = '#06283D'

root = Tk()
root.title("Student Registration Form")
root.geometry("1250x750+210+100")
root.config(bg=background)

# Create Excel file if it doesn't exist
file = pathlib.Path('student_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Gender"
    sheet['D1'] = "Class"
    sheet['E1'] = "DOB"
    sheet['F1'] = "Country"
    sheet['G1'] = "Email"
    sheet['H1'] = "Father's Name"
    file.save('student_data.xlsx')

def clear_fields():
    name_entry.delete(0, END)
    gender_combobox.set('')
    class_entry.delete(0, END)
    dob_entry.delete(0, END)
    country_entry.delete(0, END)
    email_entry.delete(0, END)
    father_name_entry.delete(0, END)

def submit_form():
    name = name_entry.get()
    gender = gender_combobox.get()
    class_name = class_entry.get()
    dob = dob_entry.get()
    country = country_entry.get()
    email = email_entry.get()
    father_name = father_name_entry.get()

    if not name or not gender or not class_name or not dob or not country or not email or not father_name:
        messagebox.showerror("Input Error", "All fields are required.")
        return

    try:
        file = load_workbook('student_data.xlsx')
        sheet = file.active
        row = sheet.max_row + 1

        sheet.cell(row, 1, row - 1)
        sheet.cell(row, 2, name)
        sheet.cell(row, 3, gender)
        sheet.cell(row, 4, class_name)
        sheet.cell(row, 5, dob)
        sheet.cell(row, 6, country)
        sheet.cell(row, 7, email)
        sheet.cell(row, 8, father_name)

        file.save('student_data.xlsx')

        messagebox.showinfo("Registration Successful", f"Student {name} has been registered successfully!")
        clear_fields()
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while saving data: {e}")

def display_data():
    try:
        file = load_workbook('student_data.xlsx')
        sheet = file.active

        display_window = Toplevel(root)
        display_window.title("Student Details")
        display_window.geometry("1000x1000")

        tree = ttk.Treeview(display_window, columns=("RegNo", "Name", "Gender", "Class", "DOB", "Country", "Email", "FatherName"), show='headings')
        tree.heading("RegNo", text="Registration No.")
        tree.heading("Name", text="Name")
        tree.heading("Gender", text="Gender")
        tree.heading("Class", text="Class")
        tree.heading("DOB", text="DOB")
        tree.heading("Country", text="Country")
        tree.heading("Email", text="Email")
        tree.heading("FatherName", text="Father's Name")
        tree.pack(fill=BOTH, expand=1)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            tree.insert("", END, values=row)

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while displaying data: {e}")

# Top frames
Label(root, text="Email: xxxxxxxxxxxxx@gmail.com", width=10, height=3, bg="light cyan", anchor='e', font=('Arial', 14, 'bold')).pack(side=TOP, fill=X)
Label(root, text="Student Registration", width=10, height=3, bg="#f0687c", anchor='center', font=('Arial', 24, 'bold')).pack(side=TOP, fill=X)

# Registration Form Frame
form_frame = Frame(root, bg=framebg, bd=2, relief=RIDGE)
form_frame.place(x=100, y=100, width=1050, height=500)

# Title inside frame
Label(form_frame, text="Registration Form", font=('Arial', 20, 'bold'), bg=framebg, fg=framefg).place(x=400, y=20)

# Input Fields
Label(form_frame, text="Name:", font=('Arial', 15), bg=framebg).place(x=50, y=80)
name_entry = Entry(form_frame, width=40, bd=2)
name_entry.place(x=250, y=80)

Label(form_frame, text="Gender:", font=('Arial', 15), bg=framebg).place(x=50, y=130)
gender_combobox = Combobox(form_frame, values=['Male', 'Female'], state='readonly', width=37)
gender_combobox.place(x=250, y=130)

Label(form_frame, text="Class:", font=('Arial', 15), bg=framebg).place(x=50, y=180)
class_entry = Entry(form_frame, width=40, bd=2)
class_entry.place(x=250, y=180)

Label(form_frame, text="Date of Birth:", font=('Arial', 15), bg=framebg).place(x=50, y=230)
dob_entry = Entry(form_frame, width=40, bd=2)
dob_entry.place(x=250, y=230)

Label(form_frame, text="Country:", font=('Arial', 15), bg=framebg).place(x=50, y=280)
country_entry = Entry(form_frame, width=40, bd=2)
country_entry.place(x=250, y=280)

Label(form_frame, text="Email:", font=('Arial', 15), bg=framebg).place(x=50, y=330)
email_entry = Entry(form_frame, width=40, bd=2)
email_entry.place(x=250, y=330)

Label(form_frame, text="Father's Name:", font=('Arial', 15), bg=framebg).place(x=50, y=380)
father_name_entry = Entry(form_frame, width=40, bd=2)
father_name_entry.place(x=250, y=380)

# Buttons
Button(form_frame, text="Submit", command=submit_form, width=15, height=2, bg="light green", font=('Arial', 12)).place(x=250, y=430)
Button(form_frame, text="Clear", command=clear_fields, width=15, height=2, bg="light blue", font=('Arial', 12)).place(x=450, y=430)
Button(form_frame, text="Display", command=display_data, width=15, height=2, bg="light yellow", font=('Arial', 12)).place(x=650, y=430)

root.mainloop()
